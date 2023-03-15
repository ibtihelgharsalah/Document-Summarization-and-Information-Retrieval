import os # used to read and write files to the server's file system
import pathlib # to get the file extension
import docx # deals with the word files
import PyPDF2 # deals with pdf files
from openpyxl import load_workbook # deals with xlsx files
import csv # deals with csv files
from utils.allowed_extention import allowed_file

# Function to preprocess the uploaded files
def extract_text(upload_folder, filenames):   
    # Create an empty list to store the document texts
    doc_texts = {}
    # Loop over the uploaded files and extract text within each file 
    for filename in filenames:  
        if allowed_file(filename):
            # Check the extension of the file
            ex = pathlib.Path(filename).suffix 
            if ex == ".pdf":
                # Read the contents of the PDF file in binary mode
                with open(os.path.join(upload_folder, filename), 'rb') as f:
                    pdf = PyPDF2.PdfReader(f)
                    # Extract the text from the PDF document
                    text = '\n'.join([pdf.pages[i].extract_text() for i in range(len(pdf.pages))])
                    doc_texts[filename] = text

            elif ex == ".docx":
                # Load the Word document
                doc = docx.Document(os.path.join(upload_folder, filename))
                # Extract the text from the document
                text = '\n'.join([para.text for para in doc.paragraphs])
                doc_texts[filename] = text

            elif ex in ['.xls', '.xlsx']:
                # Load the entire workbook
                wb = load_workbook(os.path.join(upload_folder, filename), 'r')
                # Loop over each sheet in the workbook
                for worksheet in wb.worksheets:
                    # Extract the text from each cell in the sheet
                    xlsx_text = ""
                    for row in worksheet.iter_rows(values_only=True):
                        for cell in row:
                            if cell is not None:
                                xlsx_text += str(cell) + " "
                doc_texts[filename] = xlsx_text

            elif ex == ".csv":
                # Open the CSV file and read the contents
                with open(os.path.join(upload_folder, filename), 'r') as f:
                    reader = csv.reader(f)
                    # Extract the text from each row in the CSV file
                    csv_text = ""
                    for row in reader:
                        csv_text += ",".join(row) + " "
                    doc_texts[filename] = csv_text

            elif ex == ".txt": #(text file)
                # Read the contents of the file
                with open(os.path.join(upload_folder, filename), 'r') as f:
                    text = f.read()
                    doc_texts[filename] = text
        else : print ("file not allowed")
    return doc_texts